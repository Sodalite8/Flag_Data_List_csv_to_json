import json
import openpyxl as op


INPUT_XLSX_PATH = "./flags.xlsx"
OUTPUT_JSON_PATH = "./flags.json"
ID_HEADER = "ID"
INT_DATA_HEADERS = ["level", "category"]
STRING_DATA_HEADERS = ["name", "period"]
BOOL_DATA_HEADERS = ["available"]


# シートから目的の列rowをリスト化して取り出す
def get_list1d(row):
    return [cell.value for cell in row]


# シートから目的の表list_generatorをリスト化して取り出す
def get_list2d(list_generator):
    return [[cell.value for cell in row] for row in list_generator]


# main
def main():
    # 目的のワークシートを取り出す
    wb = op.load_workbook(INPUT_XLSX_PATH)
    ws = wb.worksheets[0]

    # ヘッダーとデータの表を取り出す
    flag_header_list = get_list1d(ws[1])
    flag_data_list = get_list2d(ws.iter_rows(min_row=2, min_col=1))

    # JSONファイルに書き出すためにデータを要素が辞書型の配列に変形
    output_list = [{} for _ in range(len(flag_data_list))]
    for i in range(len(flag_data_list)):
        for j in range(len(flag_header_list)):
            current_header = flag_header_list[j]
            current_cell_value = flag_data_list[i][j]

            # idが列番号と一致しない場合、idを列番号に変える
            if current_header == ID_HEADER and current_cell_value != i:
                output_list[i][current_header] = i

            # セルが空の場合、-1や""に変える
            elif current_cell_value is None:
                if current_header in INT_DATA_HEADERS:
                    output_list[i][current_header] = -1
                elif current_header in STRING_DATA_HEADERS:
                    output_list[i][current_header] = ""
                elif current_header in BOOL_DATA_HEADERS:
                    output_list[i][current_header] = False

            # 異常がない場合、そのまま代入
            else:
                output_list[i][current_header] = current_cell_value

    # JSONファイルに書き出す
    with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as output_json:
        json.dump(output_list, output_json, indent=4, ensure_ascii=False)

    print("Convert is completed.")


if __name__ == "__main__":
    main()
